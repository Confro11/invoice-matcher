"""
Invoice Matcher - FastAPI backend
Párování faktur s platbami z platebního terminálu.
"""

import io
import tempfile
import os
from typing import List, Optional

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, File, UploadFile, HTTPException, Query
from starlette.background import BackgroundTask
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

app = FastAPI(title="Invoice Matcher")

# ── Statické soubory ──────────────────────────────────────────────────────────
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
async def root():
    return FileResponse("static/index.html")


# ── Pomocné funkce ────────────────────────────────────────────────────────────

def load_invoices(file_bytes: bytes, filename: str) -> pd.DataFrame:
    buf = io.BytesIO(file_bytes)
    df = pd.read_excel(buf, engine="calamine")
    df.columns = df.columns.str.strip()
    return df


def load_payments(file_bytes: bytes, filename: str) -> pd.DataFrame:
    buf = io.BytesIO(file_bytes)
    tables = pd.read_html(buf, encoding="cp1250", header=2)
    return tables[0]


def match(inv_raw: pd.DataFrame, pay_raw: pd.DataFrame):
    # Čištění faktur
    inv = inv_raw.dropna(subset=["Číslo"]).copy()
    inv["Vystaveno_dt"] = pd.to_datetime(inv["Vystaveno"], dayfirst=True, errors="coerce")
    inv["Castka"] = pd.to_numeric(inv["Celkem k úhradě"], errors="coerce")
    inv = inv[inv["Castka"] > 0].copy()
    inv["Stav_parovani"] = "Nezaplaceno"
    inv["Prirazena_platba"] = ""
    inv["Kód autorizace"] = ""

    # Čištění plateb
    pay = pay_raw.dropna(subset=["Datum a hodina transakce"]).copy()
    pay["Datum_dt"] = pd.to_datetime(pay["Datum a hodina transakce"], dayfirst=True, errors="coerce")
    pay["Castka"] = pd.to_numeric(pay["Částka transakce brutto"], errors="coerce")
    pay = pay[pay["Castka"] > 0].copy()
    pay["Stav_parovani"] = "Nepřiřaditelná"
    pay["Prirazena_faktura"] = ""

    used_payments = set()

    for i, inv_row in inv.iterrows():
        if pd.isna(inv_row["Vystaveno_dt"]) or pd.isna(inv_row["Castka"]):
            continue

        castka = inv_row["Castka"]
        datum_fakt = inv_row["Vystaveno_dt"].date()

        for j, pay_row in pay.iterrows():
            if j in used_payments:
                continue
            if pd.isna(pay_row["Datum_dt"]):
                continue
            if abs(pay_row["Castka"] - castka) > 0.01:
                continue
            if pay_row["Datum_dt"].date() != datum_fakt:
                continue

            inv.at[i, "Stav_parovani"] = "Zaplaceno"
            inv.at[i, "Prirazena_platba"] = str(pay_row.get("Datum a hodina transakce", ""))
            inv.at[i, "Kód autorizace"] = (
                str(int(pay_row["Kód autorizace"]))
                if pd.notna(pay_row.get("Kód autorizace"))
                else ""
            )

            pay.at[j, "Stav_parovani"] = "Přiřazená"
            pay.at[j, "Prirazena_faktura"] = inv_row["Číslo"]
            used_payments.add(j)
            break

    return inv, pay


def build_excel(inv: pd.DataFrame, pay: pd.DataFrame, sheet: str = "both") -> bytes:
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")

    def write_sheet(ws, df, status_col, color_map):
        headers = list(df.columns)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = YELLOW
            cell.font = Font(bold=True)
        for _, row in df.iterrows():
            ws.append(list(row))
            status = row.get(status_col, "")
            fill = color_map.get(status, PatternFill())
            last_row = ws.max_row
            for col in range(1, len(headers) + 1):
                ws.cell(last_row, col).fill = fill
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), *[len(str(v)) if pd.notna(v) else 0 for v in df[col]])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    wb = openpyxl.Workbook()
    first_sheet = True

    if sheet in ("invoices", "both"):
        ws_inv = wb.active if first_sheet else wb.create_sheet("Faktury")
        ws_inv.title = "Faktury"
        first_sheet = False

        inv_out = inv[
            ["Číslo", "Vystaveno", "Odběratel", "Celkem k úhradě", "Celkem s DPH",
             "Měna", "Stav_parovani", "Prirazena_platba", "Kód autorizace"]
        ].copy()
        inv_out.columns = [
            "Číslo faktury", "Vystaveno", "Odběratel", "K úhradě", "Celkem s DPH",
            "Měna", "Stav", "Přiřazená platba", "Kód autorizace"
        ]
        write_sheet(ws_inv, inv_out, "Stav", {"Zaplaceno": GREEN, "Nezaplaceno": RED})

    if sheet in ("payments", "both"):
        # Sestavení lookup tabulky faktur: číslo → (odběratel, částka)
        inv_lookup = {}
        for _, row in inv.iterrows():
            cislo = str(row["Číslo"])
            inv_lookup[cislo] = {
                "odberatel": str(row.get("Odběratel", "")) if pd.notna(row.get("Odběratel")) else "",
                "castka": row.get("Celkem k úhradě", ""),
            }

        # Sestavení výstupu plateb s obohacenými sloupci
        rows = []
        for _, row in pay.iterrows():
            cislo_fakt = str(row.get("Prirazena_faktura", ""))
            info = inv_lookup.get(cislo_fakt, {"odberatel": "", "castka": ""})
            rows.append({
                "Datum a čas": row.get("Datum a hodina transakce", ""),
                "Částka": row.get("Částka transakce brutto", ""),
                "Kód autorizace": (
                    str(int(row["Kód autorizace"]))
                    if pd.notna(row.get("Kód autorizace"))
                    else ""
                ),
                "Stav": row.get("Stav_parovani", ""),
                "Číslo faktury": cislo_fakt if cislo_fakt else "",
                "Odběratel": info["odberatel"],
                "Částka faktury": info["castka"] if cislo_fakt else "",
            })
        pay_out = pd.DataFrame(rows)

        ws_pay = wb.active if first_sheet else wb.create_sheet("Platby")
        ws_pay.title = "Platby"
        write_sheet(ws_pay, pay_out, "Stav", {"Přiřazená": GREEN, "Nepřiřaditelná": RED})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Endpoint: párování ────────────────────────────────────────────────────────

# Temporary storage for the last matched results
_last_inv: Optional[pd.DataFrame] = None
_last_pay: Optional[pd.DataFrame] = None


@app.post("/match")
async def match_files(
    invoices: List[UploadFile] = File(...),
    payments: List[UploadFile] = File(...),
):
    global _last_inv, _last_pay

    # Načtení a sloučení všech faktur
    inv_frames = []
    for f in invoices:
        data = await f.read()
        try:
            inv_frames.append(load_invoices(data, f.filename))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Chyba při čtení faktury '{f.filename}': {e}")

    # Načtení a sloučení všech plateb
    pay_frames = []
    for f in payments:
        data = await f.read()
        try:
            pay_frames.append(load_payments(data, f.filename))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Chyba při čtení plateb '{f.filename}': {e}")

    if not inv_frames or not pay_frames:
        raise HTTPException(status_code=400, detail="Je třeba nahrát alespoň jeden soubor každého typu.")

    inv_all = pd.concat(inv_frames, ignore_index=True)
    pay_all = pd.concat(pay_frames, ignore_index=True)

    try:
        inv_result, pay_result = match(inv_all, pay_all)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Chyba při párování: {e}")

    _last_inv = inv_result
    _last_pay = pay_result

    # Sestavení JSON odpovědi
    def safe(val):
        if pd.isna(val) if not isinstance(val, str) else False:
            return None
        return str(val) if not isinstance(val, (int, float, str)) else val

    invoices_out = []
    for _, row in inv_result.iterrows():
        invoices_out.append({
            "cislo": safe(row.get("Číslo", "")),
            "vystaveno": safe(row.get("Vystaveno", "")),
            "odberatel": safe(row.get("Odběratel", "")),
            "castka": safe(row.get("Celkem k úhradě", "")),
            "stav": safe(row.get("Stav_parovani", "")),
            "prirazena_platba": safe(row.get("Prirazena_platba", "")),
            "kod_autorizace": safe(row.get("Kód autorizace", "")),
        })

    payments_out = []
    for _, row in pay_result.iterrows():
        payments_out.append({
            "datum": safe(row.get("Datum a hodina transakce", "")),
            "castka": safe(row.get("Částka transakce brutto", "")),
            "kod_autorizace": (str(int(v)) if isinstance(v, float) and not pd.isna(v) else str(v) if pd.notna(v) else "") if (v := row.get("Kód autorizace", "")) != "" else "",
            "stav": safe(row.get("Stav_parovani", "")),
            "prirazena_faktura": safe(row.get("Prirazena_faktura", "")),
        })

    # Statistiky
    stats = {
        "faktury_celkem": len(inv_result),
        "zaplaceno": int((inv_result["Stav_parovani"] == "Zaplaceno").sum()),
        "nezaplaceno": int((inv_result["Stav_parovani"] == "Nezaplaceno").sum()),
        "platby_celkem": len(pay_result),
        "prirazene": int((pay_result["Stav_parovani"] == "Přiřazená").sum()),
        "nepriraditelne": int((pay_result["Stav_parovani"] == "Nepřiřaditelná").sum()),
    }

    return JSONResponse({
        "stats": stats,
        "invoices": invoices_out,
        "payments": payments_out,
    })


@app.get("/download")
async def download_xlsx(
    sheet: str = Query(default="both", pattern="^(invoices|payments|both)$")
):
    global _last_inv, _last_pay
    if _last_inv is None or _last_pay is None:
        raise HTTPException(status_code=404, detail="Nejprve proveďte párování.")

    xlsx_bytes = build_excel(_last_inv, _last_pay, sheet=sheet)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(xlsx_bytes)
        tmp_path = tmp.name

    name_map = {"invoices": "faktury", "payments": "platby", "both": "parovani_faktur"}
    filename = f"{name_map[sheet]}.xlsx"

    return FileResponse(
        tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        background=BackgroundTask(os.unlink, tmp_path),
    )
